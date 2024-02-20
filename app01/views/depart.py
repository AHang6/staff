from django.shortcuts import render, redirect

from app01.models import Department
from app01.utils.pagination import Pagination
from openpyxl import load_workbook



def depart_list(request):
    """  部门列表  """
    data_list = Department.objects.all()

    # for i in range(1, 200):
    #     Department.objects.create(title="技术部")
    page_obj = Pagination(request, data_list)

    context = {
        'data_list': page_obj.page_queryset,
        'page_string': page_obj.html()
    }

    return render(request, 'depart_list.html', context)


def depart_add(request):
    """  添加用户  """
    if request.method == "GET":
        return render(request, 'depart_add.html')

    title = request.POST.get("title")

    Department.objects.create(title=title)
    # 重定向
    return redirect("/depart/list/")


def depart_delete(request):
    """  删除用户  """
    nid = request.GET.get('nid')

    Department.objects.filter(id=nid).delete()

    return redirect('/depart/list/')


def depart_edit(request):
    """   修改部门   """
    nid = request.GET.get("nid")
    if request.method == 'GET':
        depart_info = Department.objects.filter(id=nid).all().first()
        return render(request, "depart_edit.html", {'depart_title': depart_info.title})

    depart_title = request.POST.get('title')
    Department.objects.filter(id=nid).update(title=depart_title)
    return redirect('/depart/list/')


def depart_multi(request):
    file_obj = request.FILES.get('avatar')

    print(type(file_obj))

    wb = load_workbook(file_obj)
    sheet = wb.worksheets[0]

    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        print(text)
        exists = Department.objects.filter(title=text).exists()
        if not exists:
            Department.objects.create(title=text)

    return redirect('/depart/list/')

